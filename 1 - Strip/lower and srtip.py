import openpyxl

def strip_excel_data(input_file, output_file):
    # Open the input Excel file
    input_workbook = openpyxl.load_workbook(input_file)
    input_sheet = input_workbook.active

    # Create a new workbook for output
    output_workbook = openpyxl.Workbook()
    output_sheet = output_workbook.active

    # Strip whitespace from each cell and write to the output sheet
    for row in input_sheet.iter_rows(values_only=True):
        stripped_row = [str(cell).strip() if cell is not None else None for cell in row]
        output_sheet.append(stripped_row)

    # Save the output workbook
    output_workbook.save(output_file)
    print(f"Stripped data written to {output_file}")


input_file_path = "Sample Data.xlsx"
output_file_path = "Sample Data Cleaned.xlsx"

strip_excel_data(input_file_path, output_file_path)
